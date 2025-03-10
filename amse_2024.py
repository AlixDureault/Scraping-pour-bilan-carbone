import requests
from bs4 import BeautifulSoup
from pprint import pprint
from urllib.parse import urljoin
from openpyxl import load_workbook
import os
from datetime import datetime
import locale

try:
    locale.setlocale(locale.LC_TIME, "fr_FR.utf8")
except locale.Error:
    locale.setlocale(locale.LC_TIME, "French_France.1252")

# Toutes les fonctions

def seminar_list_function (soup):
    seminar_list_div = soup.find('div', id='ancre-content')
    if seminar_list_div is None:
        return []
    seminar_list = seminar_list_div.find_all('div', class_="node node-events node-teaser events-teaser clearfix")
    return(seminar_list)

def next_page_url_function (soup) :
    next_page_link_div = soup.find('li', class_="pager__item").find('a')
    if next_page_link_div == None :
        return None
    else :
        next_page_link_half = next_page_link_div['href']
        next_page_full_url = urljoin(url_AMSE,next_page_link_half)
        return(next_page_full_url)

def get_date_hour_function (seminar) :
    date_div = seminar.find('span', class_='date-display-single')
    if date_div != None :
        date_hour = date_div.text.split("|")
        date = date_hour[0]
        hour = date_hour[1]
    else :
        date_div = seminar.find('span', class_='date-display-range')
        hour = 'Plusieurs jours'
        date = date_div.text
    return date, hour

def name_intervenant_function (seminar) :
    person_name_div = seminar.find('div', class_='col-md-9').find('div', class_="field-item even").find("a")
    person_name = person_name_div.text
    return (person_name)

def event_url_function (seminar) :
    person_name_div = seminar.find('div', class_='col-md-9').find('div', class_="field-item even").find("a")
    event_link = person_name_div['href']
    event_full_url = urljoin(url_AMSE, event_link)
    return (event_full_url)

def intervenant_origin_university_function (seminar):
    person_origin_university_div = seminar.find('div', class_='col-md-9').find('div', class_="field field-name-field-event-subtitle field-type-text field-label-hidden")
    person_origin_university = person_origin_university_div.text.strip()
    return(person_origin_university)

def seminar_title_function(seminar) :
    seminar_title_div = seminar.find('div', class_='col-md-9').find_all('div', class_="field field-name-field-event-paper-title field-type-text field-label-hidden")
    seminar_title_list = [div.text.strip() for div in seminar_title_div]
    delimiter = ","
    seminar_title_str = delimiter.join(seminar_title_list)
    return (seminar_title_str)

def tag1_function (seminar) :
    seminar_tags = seminar.find('div', class_='col-md-9').find('ul', 'tags-events-child')
    seminar_tag1_div = seminar_tags.find('li', "parent-term")
    seminar_tag1 = seminar_tag1_div.text
    return (seminar_tag1)

def tag2_function (seminar) :
    seminar_tags = seminar.find('div', class_='col-md-9').find('ul', 'tags-events-child')
    seminar_tag2_div = seminar_tags.find_all('li', "child-term")
    seminar_tag2_list = [li.text for li in seminar_tag2_div]
    delimiter = ","
    seminar_tag2_str = delimiter.join(seminar_tag2_list)
    return (seminar_tag2_str)

def lieu_function (event_full_url) :
    response2 = requests.get(event_full_url)
    soup2 = BeautifulSoup(response2.text, 'html.parser')

    location_div = soup2.find('div', class_='taxonomy-term vocabulary-lieux-salles').find('div', class_ ='field field-name-field-salles-adresse field-type-text-long field-label-hidden').find('div', class_='field-item even')
    location = location_div.text.strip()
    return (location)

def contacts_function (event_full_url) :
    # Ne pas appliquer aux soutenances de thèse à priori
    response2 = requests.get(event_full_url)
    soup2 = BeautifulSoup(response2.text, 'html.parser')

    contacts_name_div = soup2.find("div", class_="field field-name-field-event-contact field-type-text-long field-label-above label-text").find('div', class_="field-items")
    contacts_name = contacts_name_div.text.strip()
    return(contacts_name)

def jury_names_universities_function (event_full_url):
    # N'appliquer qu'aux soutenances de stage
    response2 = requests.get(event_full_url)
    soup2 = BeautifulSoup(response2.text, 'html.parser')

    jury_names_div = soup2.find("div", class_="field field-name-field-event-contenu-import field-type-text-long field-label-above label-text").find('div', class_="field-items").find_all('li')
    jury_names_list = [a.text for a in jury_names_div]
    delimiter = "   /   "
    jury_names_str = delimiter.join(jury_names_list).replace('\xa0', '')
    return(jury_names_str)

def get_info_from_page(url):
    try :
        response = requests.get(url)
        response.raise_for_status()
    except requests.RequestException as e :
        print(f"Erreur lors de l'accès à {url} : {e}")
        return [], None
    soup = BeautifulSoup(response.text, 'html.parser')
    seminar_list = seminar_list_function(soup)
    next_page_url = next_page_url_function(soup)
    return (seminar_list, next_page_url)

def surf_all_pages(url, start_date, end_date) :
    info_page_1 = get_info_from_page (url)
    seminar_total_list = []
    next_page_url = info_page_1[1]

    for seminar in info_page_1[0] :
            seminar_date_raw = get_date_hour_function(seminar)[0]
            seminar_date = convertir_date_scraping(seminar_date_raw)

            if start_date <= seminar_date <= end_date:
                seminar_total_list.append(seminar)

    if next_page_url != None :
        seminar_total_list.extend(surf_all_pages(next_page_url, start_date, end_date))
    
    return (seminar_total_list)

def event_cancellation_function (seminar) :
    seminar_tags = seminar.find('div', class_='col-md-9').find('ul', class_='cancel-tag-ul')
    return (seminar_tags)

def event_online_function (seminar) :
    seminar_online_tag = seminar.find('div', class_='col-md-9').find('ul', 'online-tag-ul')
    return(seminar_online_tag)

def est_fichier_ouvert(fichier):
    try:
        with open(fichier, 'r+'):
            return False
    except IOError:
        return True

def convertir_date_scraping(date_str):
    """Convertit une date sous forme de chaîne en objet datetime."""
    
    # Séparer les composants de la date ("Jeudi 1 février 2024")
    parties = date_str.split("|")[0].split()  # ["Jeudi", "1", "février", "2024"]
    
    jour = parties[1]  # "1"
    mois = mois_francais[parties[2].lower()]  # "février" -> "02"
    annee = parties[3]  # "2024"
    
    # Reformater en "YYYY-MM-DD" et convertir en datetime
    date_formatee = f"{annee}-{mois}-{jour.zfill(2)}"  # "2024-02-01"
    return datetime.strptime(date_formatee, "%Y-%m-%d")

mois_francais = {
    "janvier": "01",
    "février": "02",
    "mars": "03",
    "avril": "04",
    "mai": "05",
    "juin": "06",
    "juillet": "07",
    "août": "08",
    "septembre": "09",
    "octobre": "10",
    "novembre": "11",
    "décembre": "12"
}

# Main

with requests.Session() as session :
 
    fichier = "AMSE_2024.xlsx"
    nom_feuilles = ["Inputs", "Séminaires", "Soutenances de thèses","Autres séminaires", "Evènements en distanciel", "Evènements annulés"]

    if est_fichier_ouvert(fichier):
        print(f"Le fichier {fichier} est ouvert, veuillez le fermer avant de lancer le code")
    else :
        try :
            wb = load_workbook(fichier)
        except Exception as e :
            print(f"Erreur lors de l'ouverture du fichier : {e}")
            exit()

        feuilles = {}
        for nom in nom_feuilles :
            if nom in wb.sheetnames:
                feuilles[nom] = wb[nom]
            else:
                print(f"La feuille '{nom}' est manquante.")
                exit()

        url_AMSE = feuilles["Inputs"]["C2"].value
        url_2024 = feuilles["Inputs"]["C3"].value
        start_date = feuilles["Inputs"]["C4"].value
        end_date = feuilles["Inputs"]["C5"].value

        seminar_list_div = surf_all_pages (url_2024, start_date, end_date)
        info_seminar_list = []
        info_soutenances_list = []
        info_others_list = []
        cancelled_events = []
        online_events = []

        for seminar in seminar_list_div :

            if event_cancellation_function(seminar) == None :

                if event_online_function(seminar) == None :

                    tag1 = tag1_function(seminar)

                    if tag1 != 'grand public' and tag1 != 'conférences/workshops' and tag1 != 'enseignement' and tag1 != 'soutenances de thèse' :
                        event_url = event_url_function(seminar)
                        info_seminar = [tag1, tag2_function(seminar), seminar_title_function(seminar), get_date_hour_function(seminar)[0], get_date_hour_function(seminar)[1], lieu_function(event_url), name_intervenant_function(seminar), intervenant_origin_university_function(seminar), contacts_function(event_url), event_url]
                        info_seminar_list.append(info_seminar)

                    elif tag1 == 'soutenances de thèse' :
                        event_url = event_url_function(seminar)
                        info_seminar = [tag1, name_intervenant_function(seminar), seminar_title_function(seminar), intervenant_origin_university_function(seminar), jury_names_universities_function(event_url), get_date_hour_function(seminar)[0], get_date_hour_function(seminar)[1], lieu_function(event_url), event_url]
                        info_soutenances_list.append(info_seminar)

                    else :
                        event_url = event_url_function(seminar)
                        info_seminar = [tag1, tag2_function(seminar), get_date_hour_function(seminar)[0], get_date_hour_function(seminar)[1], name_intervenant_function(seminar), event_url]
                        info_others_list.append(info_seminar)

                else :
                    event_url = event_url_function(seminar)
                    info_seminar = [tag1, tag2_function(seminar), seminar_title_function(seminar), get_date_hour_function(seminar)[0], get_date_hour_function(seminar)[1], name_intervenant_function(seminar), intervenant_origin_university_function(seminar), contacts_function(event_url), event_url]
                    online_events.append(info_seminar)

            else :
                event_url = event_url_function(seminar)
                info_seminar = [tag1, tag2_function(seminar), seminar_title_function(seminar), get_date_hour_function(seminar)[0], get_date_hour_function(seminar)[1], name_intervenant_function(seminar), intervenant_origin_university_function(seminar), contacts_function(event_url), event_url]
                cancelled_events.append(info_seminar)

        start_row = 3
        start_col = 2

        for i, row in enumerate(info_seminar_list, start=start_row) :
            for j, value in enumerate(row, start=start_col) :
                feuilles["Séminaires"].cell(row = i, column = j, value = value)

        for i, row in enumerate(info_soutenances_list, start=start_row) :
            for j, value in enumerate(row, start=start_col) :
                feuilles["Soutenances de thèses"].cell(row = i, column = j, value = value)

        for i, row in enumerate(online_events, start=start_row) :
            for j, value in enumerate(row, start=start_col) :
                feuilles["Evènements en distanciel"].cell(row = i, column = j, value = value)

        for i, row in enumerate(cancelled_events, start=start_row) :
            for j, value in enumerate(row, start=start_col) :
                feuilles["Evènements annulés"].cell(row = i, column = j, value = value)

        for i, row in enumerate(info_others_list, start=start_row) :
            for j, value in enumerate(row, start=start_col) :
                feuilles["Autres séminaires"].cell(row = i, column = j, value = value)
        
        wb.save(fichier)

        os.startfile(fichier)

        print("Données ajoutées avec succès !")